#!/usr/bin/env python3
"""Programmatic QA for an .xlsx — no LibreOffice required.

Function: verify formulas by evaluating the workbook in pure Python (the
          `formulas` package) and reporting any cell that resolves to an Excel
          error (#DIV/0!, #REF!, ...). Falls back to an openpyxl static scan
          for already-materialized error values if `formulas` isn't installed.
Usage:    python scripts/qa_check.py <file.xlsx>
Output:   a human-readable report; exits non-zero if any error cell is found
          (or if formulas can't be evaluated and it can't be ruled out).
"""
import sys
import warnings

from openpyxl import load_workbook

ERROR_PREFIX = "#"
KNOWN_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")


def evaluate_with_formulas(path):
    """Return list of (cell, error) by computing the whole workbook. None if lib missing."""
    try:
        import formulas
    except ImportError:
        return None
    warnings.filterwarnings("ignore")
    sol = formulas.ExcelModel().loads(path).finish().calculate()
    errs = []
    for cell, rng in sol.items():
        try:
            val = rng.value[0, 0]
        except Exception:
            continue
        if str(val).startswith(ERROR_PREFIX):
            errs.append((cell.split("!")[-1], str(val)))
    return errs


def static_scan(path):
    """Scan already-computed cell values for error strings (only catches materialized errors)."""
    wb = load_workbook(path, data_only=True)
    hits = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value in KNOWN_ERRORS:
                    hits.append((f"{ws.title}!{cell.coordinate}", cell.value))
    return hits


def has_formulas(path):
    wb = load_workbook(path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    return True
    return False


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/qa_check.py <file.xlsx>", file=sys.stderr)
        sys.exit(2)
    path = sys.argv[1]

    errs = evaluate_with_formulas(path)
    if errs is None:
        print("NOTE: the `formulas` package is not installed — falling back to a static scan.")
        print("      For real formula verification without LibreOffice: pip install formulas")
        errs = static_scan(path)
        if not errs and has_formulas(path):
            print("WARN: this file has formulas but they could not be evaluated "
                  "(install `formulas` to verify, or the cached values may be absent).")
            print("RESULT: INCONCLUSIVE")
            sys.exit(1)

    if errs:
        print(f"error cells ({len(errs)}):")
        for cell, val in errs:
            print(f"  {cell}: {val}")
        print("RESULT: FAIL")
        sys.exit(1)

    print("error cells: none")
    print("RESULT: PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
